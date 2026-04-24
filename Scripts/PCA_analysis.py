#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PCA + Cronbach's Alpha + Composite Analysis
Brazilian Legal Complexity Features — Diachronic Study

Outputs:
  - Console: eigenvalues, loadings, alpha, PC1 scores
  - Scree plot (matplotlib)
  - PC trends plot (matplotlib)
  - PC1 year-level plot (matplotlib)
  - complexity_composite_analysis.xlsx  (10 sheets)

Requirements: pip install numpy pandas openpyxl matplotlib
"""

import json
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from matplotlib.lines import Line2D

# ── Config ────────────────────────────────────────────────────────────────────
INPUT_JSON   = 'complexity_analysis_data.json'
OUTPUT_XLS   = 'complexity_composite_analysis.xlsx'
METRIC_GROUP = 'per_sentence'          # or 'per_100_tokens'

BINS = [
    ("1943–1962", 1943, 1962),
    ("1964–1986", 1964, 1986),
    ("1988–2016", 1988, 2016),
    ("2017–2022", 2017, 2022),
]

FEATURES = OrderedDict([
    ('word_order_inversion',  'Word Order Inversion (VS)'),
    ('nominalization',        'Nominalizations'),
    ('subordination',         'Subordinate Clauses'),
    ('passive',               'Passive Constructions'),
    ('gerunds_participles',   'Gerunds & Participles'),
    ('relative_pronouns',     'Relative Pronouns & Conj.'),
    ('appositions',           'Appositions & Parentheticals'),
    ('prepositional_phrases', 'Prepositional Phrases'),
    ('complex_verb_forms',    'Complex Verb Forms'),
])

# ── Excel styles ──────────────────────────────────────────────────────────────
DARK_FILL    = PatternFill("solid", fgColor="1F3864")
MED_FILL     = PatternFill("solid", fgColor="2E75B6")
ALT_FILL     = PatternFill("solid", fgColor="DCE6F1")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
GREEN_FILL   = PatternFill("solid", fgColor="70AD47")
LTGREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
RED_FILL     = PatternFill("solid", fgColor="FCE4D6")
TITLE_FONT   = Font(bold=True, color="FFFFFF", size=12)
HDR_FONT     = Font(bold=True, color="FFFFFF", size=10)
NORMAL_FONT  = Font(size=10)
BOLD_FONT    = Font(bold=True, size=10)
CENTER       = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT         = Alignment(horizontal='left',   vertical='center', wrap_text=True)
THIN         = Side(style='thin', color='BFBFBF')
BORD         = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── Data loading ──────────────────────────────────────────────────────────────

def assign_period(year):
    for label, s, e in BINS:
        if s <= year <= e:
            return label
    return 'Outside bins'


def load_year_matrix(path, features, metric_group):
    """
    Returns:
      years         : list[int]
      X             : ndarray  (n_years, n_features)
      period_labels : list[str]
    """
    with open(path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    years, rows, plabels = [], [], []

    for y_str in sorted(data.keys(), key=int):
        rec = data[y_str]
        mg  = rec.get(metric_group, {})
        row = []
        missing = False
        for feat in features.keys():
            if feat not in mg:
                missing = True
                break
            row.append(mg[feat])
        if not missing:
            years.append(int(y_str))
            rows.append(row)
            plabels.append(assign_period(int(y_str)))

    X = np.array(rows, dtype=float)
    return years, X, plabels


def build_period_matrix(years, X_year):
    """Aggregate year-level values into period-level means."""
    labels, rows = [], []
    for label, s, e in BINS:
        idx = [i for i, y in enumerate(years) if s <= y <= e]
        if idx:
            labels.append(label)
            rows.append(X_year[idx].mean(axis=0))
    return labels, np.array(rows, dtype=float)


# ── Core statistics ───────────────────────────────────────────────────────────

def standardize(X):
    """Standardize columns to mean 0, std 1."""
    means = X.mean(axis=0)
    stds  = X.std(axis=0, ddof=1)
    stds[stds == 0] = 1.0
    return (X - means) / stds, means, stds


def pca_numpy(Xz):
    """
    PCA via eigendecomposition of the covariance matrix.
    Returns:
      eigenvalues      sorted descending
      explained_ratio  proportion of total variance per component
      components       shape (n_components, n_features)  — the loadings
    """
    cov               = np.cov(Xz, rowvar=False, ddof=1)
    eigvals, eigvecs  = np.linalg.eigh(cov)
    order             = np.argsort(eigvals)[::-1]
    eigvals           = eigvals[order]
    eigvecs           = eigvecs[:, order]
    total_var         = eigvals.sum()
    explained_ratio   = eigvals / total_var if total_var > 0 else eigvals * 0
    components        = eigvecs.T
    return eigvals, explained_ratio, components


def cronbach_alpha(X):
    """
    Cronbach's alpha.
    X: (n_observations, n_items)
    """
    k, n = X.shape[1], X.shape[0]
    if k <= 1 or n <= 1:
        return np.nan
    item_var  = X.var(axis=0, ddof=1).sum()
    total_var = X.sum(axis=1).var(ddof=1)
    if total_var == 0:
        return np.nan
    return (k / (k - 1)) * (1.0 - item_var / total_var)


def alpha_if_deleted(X, keys):
    """Return {key: alpha_without_that_feature}."""
    return {k: cronbach_alpha(np.delete(X, i, axis=1))
            for i, k in enumerate(keys)}


def backward_elimination(X, keys, feature_labels):
    """
    Greedy backward elimination:
    repeatedly remove the feature whose deletion most improves alpha.
    Returns (best_alpha, best_keys, log DataFrame).
    """
    cur_keys = list(keys)
    cur_X    = X.copy()
    cur_a    = cronbach_alpha(cur_X)

    log = [{
        'Step': 0,
        'Action': 'Initial – all features',
        "Cronbach's α": round(cur_a, 4),
        'N features': len(cur_keys),
        'Features kept': ', '.join(feature_labels[k] for k in cur_keys),
    }]

    for step in range(1, len(keys)):
        if len(cur_keys) <= 2:
            break
        candidates = {k: cronbach_alpha(np.delete(cur_X, i, axis=1))
                      for i, k in enumerate(cur_keys)}
        best_k = max(candidates, key=candidates.get)
        if candidates[best_k] <= cur_a:
            break
        best_i   = cur_keys.index(best_k)
        cur_keys.pop(best_i)
        cur_X    = np.delete(cur_X, best_i, axis=1)
        cur_a    = candidates[best_k]
        log.append({
            'Step': step,
            'Action': f"Removed: {feature_labels[best_k]}",
            "Cronbach's α": round(cur_a, 4),
            'N features': len(cur_keys),
            'Features kept': ', '.join(feature_labels[k] for k in cur_keys),
        })

    return cur_a, cur_keys, pd.DataFrame(log)


# ── Alpha interpretation ──────────────────────────────────────────────────────

def interp_alpha(a):
    if np.isnan(a):  return 'N/A'
    if a >= 0.9:     return 'Excellent  (≥ 0.90)'
    if a >= 0.8:     return 'Good       (0.80–0.89)'
    if a >= 0.7:     return 'Acceptable (0.70–0.79)'
    if a >= 0.6:     return 'Questionable (0.60–0.69)'
    if a >= 0.5:     return 'Poor       (0.50–0.59)'
    return           'Unacceptable (< 0.50)'


# ── DataFrame builders ────────────────────────────────────────────────────────

def make_alpha_summary_df(Xz_y, Xz_p, a_y, a_p):
    return pd.DataFrame([
        {
            'Level': 'Year-level',
            'N observations': Xz_y.shape[0],
            'N features': Xz_y.shape[1],
            "Cronbach's α": round(a_y, 4),
            'Interpretation': interp_alpha(a_y),
            'Note': '',
        },
        {
            'Level': 'Period-level',
            'N observations': Xz_p.shape[0],
            'N features': Xz_p.shape[1],
            "Cronbach's α": round(a_p, 4),
            'Interpretation': interp_alpha(a_p),
            'Note': 'Only 4 observations — treat as indicative only',
        },
    ])


def make_aid_df(Xz_y, Xz_p, a_y, a_p, keys, feature_labels):
    aid_y = alpha_if_deleted(Xz_y, keys)
    aid_p = alpha_if_deleted(Xz_p, keys)
    rows  = []
    for k in keys:
        rows.append({
            'Feature': feature_labels[k],
            'α if deleted (year-level)':   round(aid_y[k], 4),
            'α if deleted (period-level)': round(aid_p[k], 4),
            'Drop improves year α?':   'Yes ✓' if aid_y[k] > a_y else 'No',
            'Drop improves period α?': 'Yes ✓' if aid_p[k] > a_p else 'No',
        })
    return pd.DataFrame(rows)


def make_eigenvalue_df(vals, ratio):
    cumul = np.cumsum(ratio)
    return pd.DataFrame([{
        'Component': f'PC{i+1}',
        'Eigenvalue': round(vals[i], 4),
        'Variance explained (%)': round(ratio[i] * 100, 2),
        'Cumulative (%)': round(cumul[i] * 100, 2),
        'Kaiser (λ > 1)': 'Keep' if vals[i] > 1 else 'Drop',
    } for i in range(len(vals))])


def make_loadings_df(comps, keys, feature_labels):
    n_pc = comps.shape[0]
    rows = []
    for i, k in enumerate(keys):
        row = {'Feature': feature_labels[k]}
        for j in range(n_pc):
            row[f'PC{j+1}'] = round(float(comps[j, i]), 4)
        rows.append(row)
    return pd.DataFrame(rows)


def make_pc1_year_df(years, pc1, period_labels):
    return pd.DataFrame([{
        'Year': y,
        'Period': p,
        'PC1 score (complexity index)': round(float(s), 4),
    } for y, s, p in zip(years, pc1, period_labels)])


def make_pc123_year_df(years, pc_scores_3, period_labels):
    return pd.DataFrame([{
        'Year': y,
        'Period': p,
        'PC1 score': round(float(s1), 4),
        'PC2 score': round(float(s2), 4),
        'PC3 score': round(float(s3), 4),
    } for y, (s1, s2, s3), p in zip(years, pc_scores_3, period_labels)])


def make_period_composite_df(years, pc1, col_label='Mean PC1'):
    rows = []
    for label, s, e in BINS:
        vals = [v for y, v in zip(years, pc1) if s <= y <= e]
        if not vals:
            continue
        rows.append({
            'Period': label,
            'N years': len(vals),
            col_label: round(np.mean(vals), 4),
            'SD': round(np.std(vals, ddof=1), 4) if len(vals) > 1 else np.nan,
            'Min': round(float(np.min(vals)), 4),
            'Max': round(float(np.max(vals)), 4),
        })
    return pd.DataFrame(rows)


def make_corr_df(Xz, keys, feature_labels):
    lbls = [feature_labels[k] for k in keys]
    corr = np.corrcoef(Xz, rowvar=False)
    return pd.DataFrame(np.round(corr, 3), index=lbls, columns=lbls)


# ── Excel helpers ─────────────────────────────────────────────────────────────

def autowidth(ws, max_w=50):
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, max_w)


def write_df(ws, df, title, start_row=1):
    nc = len(df.columns)

    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row,   end_column=nc)
    tc = ws.cell(row=start_row, column=1, value=title)
    tc.font = TITLE_FONT; tc.fill = DARK_FILL; tc.alignment = CENTER

    for j, col in enumerate(df.columns, 1):
        c = ws.cell(row=start_row + 1, column=j, value=col)
        c.font = HDR_FONT; c.fill = MED_FILL
        c.alignment = CENTER; c.border = BORD

    for ri, (_, row) in enumerate(df.iterrows(), start_row + 2):
        fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL
        for j, val in enumerate(row.values, 1):
            c = ws.cell(row=ri, column=j, value=val)
            c.font = NORMAL_FONT; c.fill = fill
            c.alignment = LEFT if j == 1 else CENTER
            c.border = BORD

    ws.freeze_panes = ws.cell(row=start_row + 2, column=1)
    autowidth(ws)


def write_corr(ws, df_corr, title, start_row=1):
    nc = len(df_corr.columns) + 1

    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row,   end_column=nc)
    tc = ws.cell(row=start_row, column=1, value=title)
    tc.font = TITLE_FONT; tc.fill = DARK_FILL; tc.alignment = CENTER

    corner = ws.cell(row=start_row + 1, column=1, value='Feature')
    corner.font = HDR_FONT; corner.fill = MED_FILL; corner.alignment = CENTER

    for j, col in enumerate(df_corr.columns, 2):
        c = ws.cell(row=start_row + 1, column=j, value=col)
        c.font = HDR_FONT; c.fill = MED_FILL
        c.alignment = CENTER; c.border = BORD

    for ri, (lbl, row) in enumerate(df_corr.iterrows(), start_row + 2):
        lc = ws.cell(row=ri, column=1, value=lbl)
        lc.font = BOLD_FONT; lc.alignment = LEFT
        for j, val in enumerate(row.values, 2):
            c = ws.cell(row=ri, column=j, value=float(val))
            c.alignment = CENTER; c.font = NORMAL_FONT; c.border = BORD
            if   val >= 0.7:  c.fill = GREEN_FILL
            elif val >= 0.4:  c.fill = LTGREEN_FILL
            elif val <  0.0:  c.fill = RED_FILL
            else:             c.fill = WHITE_FILL

    ws.freeze_panes = ws.cell(row=start_row + 2, column=2)
    autowidth(ws)


# ── Plots ─────────────────────────────────────────────────────────────────────

def plot_scree(eigvals, best_keys, all_keys):
    fig, axes = plt.subplots(1, 2, figsize=(13, 5))

    pcs = np.arange(1, len(eigvals) + 1)
    axes[0].plot(pcs, eigvals, 'o-', color='#2E75B6', linewidth=2, markersize=7)
    axes[0].axhline(y=1, color='red', linestyle='--', linewidth=1.2,
                    label='Kaiser threshold (λ = 1)')
    axes[0].set_xlabel('Principal Component', fontsize=11)
    axes[0].set_ylabel('Eigenvalue', fontsize=11)
    axes[0].set_title('Scree Plot — All Features\n(Year-Level PCA)',
                      fontsize=12, fontweight='bold')
    axes[0].grid(alpha=0.3)
    axes[0].legend(fontsize=9)
    axes[0].set_xticks(pcs)

    feature_labels_list = [FEATURES[k] for k in all_keys]
    in_best = ['✓ Kept' if k in best_keys else '✗ Removed' for k in all_keys]
    colours = ['#70AD47' if k in best_keys else '#FF7F7F' for k in all_keys]
    y_pos   = np.arange(len(all_keys))
    axes[1].barh(y_pos, [1] * len(all_keys), color=colours, edgecolor='white')
    axes[1].set_yticks(y_pos)
    axes[1].set_yticklabels(feature_labels_list, fontsize=9)
    axes[1].set_xticks([])
    axes[1].set_title('Best Subset — Backward Elimination\n(Green = kept, Red = removed)',
                      fontsize=12, fontweight='bold')
    for i, (lb, col) in enumerate(zip(in_best, colours)):
        axes[1].text(0.5, i, lb, ha='center', va='center',
                     fontsize=9, color='white', fontweight='bold')

    plt.tight_layout()
    plt.savefig('scree_plot.png', dpi=150, bbox_inches='tight')
    print("  Scree plot saved to: scree_plot.png")
    plt.show()


def plot_pc_trends(years, pc1, pc2, pc3, bins=BINS, outpath="pc_trends.png"):
    df = pd.DataFrame({
        "Year": years,
        "PC1": pc1,
        "PC2": pc2,
        "PC3": pc3,
    }).sort_values("Year")

    fig, ax = plt.subplots(figsize=(12, 5))

    for label, s, e in bins:
        ax.axvspan(s, e, alpha=0.08, color="grey")
        ax.text((s + e) / 2, 0.98, label,
                transform=ax.get_xaxis_transform(),
                ha="center", va="top", fontsize=9, color="black")

    ax.plot(df["Year"], df["PC1"], marker="o", linewidth=1.8,
            markersize=4, label="PC1")
    ax.plot(df["Year"], df["PC2"], marker="o", linewidth=1.8,
            markersize=4, label="PC2")
    ax.plot(df["Year"], df["PC3"], marker="o", linewidth=1.8,
            markersize=4, label="PC3")

    ax.axhline(0, color="black", linewidth=0.8, alpha=0.6)
    ax.set_xlabel("Year", fontsize=11)
    ax.set_ylabel("PC score", fontsize=11)
    ax.set_title("PCA Scores Over Time (PC1–PC3)", fontsize=12, fontweight='bold')
    ax.grid(alpha=0.25)
    ax.legend(fontsize=10)

    plt.tight_layout()
    plt.savefig(outpath, dpi=150, bbox_inches="tight")
    print(f"  PC trends plot saved to: {outpath}")
    plt.show()


def plot_pc1_years(years, pc1, bins=BINS, outpath="pc1_over_time.png"):
    """
    Clean year-level line plot of PC1 scores over time.
    Period bins are shaded in the background.
    Uses the best-subset PC1 by default.
    """
    df = pd.DataFrame({"Year": years, "PC1": pc1}).sort_values("Year")

    # colour palette for the bins
    bin_colours = ["#DDEEFF", "#FFEEDD", "#DDFFDD", "#FFE5E5"]

    fig, ax = plt.subplots(figsize=(14, 5))

    # ── shaded period bands ───────────────────────────────────────────────────
    for (label, s, e), colour in zip(bins, bin_colours):
        ax.axvspan(s, e, alpha=0.35, color=colour, zorder=0)
        ax.text(
            (s + e) / 2, 0.97, label,
            transform=ax.get_xaxis_transform(),
            ha="center", va="top",
            fontsize=9, color="#333333",
            fontweight="bold"
        )

    # ── zero baseline ─────────────────────────────────────────────────────────
    ax.axhline(0, color="black", linewidth=0.9, linestyle="--",
               alpha=0.5, zorder=1)

    # ── main line ─────────────────────────────────────────────────────────────
    ax.plot(
        df["Year"], df["PC1"],
        color="#2E75B6",
        linewidth=1.8,
        marker="o",
        markersize=4,
        markerfacecolor="white",
        markeredgewidth=1.2,
        markeredgecolor="#2E75B6",
        zorder=2,
        label="PC1 score"
    )

    # ── period mean lines ─────────────────────────────────────────────────────
    for label, s, e in bins:
        vals = df.loc[(df["Year"] >= s) & (df["Year"] <= e), "PC1"]
        if not vals.empty:
            ax.hlines(
                vals.mean(), s, e,
                colors="red", linewidths=1.4,
                linestyles="solid", alpha=0.7,
                zorder=3
            )

    # ── labels, grid, legend ──────────────────────────────────────────────────
    ax.set_xlabel("Year", fontsize=12)
    ax.set_ylabel("PC1 Score (Syntactic Complexity)", fontsize=12)
    ax.set_title(
        "Syntactic Complexity Over Time — PC1 Score by Year\n"
        "(Best Subset: Nominalizations, Subordinate Clauses, "
        "Gerunds & Participles, Prepositional Phrases)",
        fontsize=11, fontweight="bold"
    )
    ax.grid(axis="y", alpha=0.3, zorder=0)
    ax.grid(axis="x", alpha=0.15, zorder=0)

    legend_elements = [
        Line2D([0], [0], color="#2E75B6", linewidth=1.8,
               marker="o", markersize=4,
               markerfacecolor="white", markeredgecolor="#2E75B6",
               label="PC1 score (year)"),
        Line2D([0], [0], color="red", linewidth=1.4,
               linestyle="solid", label="Period mean"),
        Line2D([0], [0], color="black", linewidth=0.9,
               linestyle="--", alpha=0.5, label="Zero baseline"),
    ]
    ax.legend(handles=legend_elements, fontsize=9, loc="upper left")

    plt.tight_layout()
    plt.savefig(outpath, dpi=150, bbox_inches="tight")
    print(f"  PC1 year-level plot saved to: {outpath}")
    plt.show()


# ── Console helpers ───────────────────────────────────────────────────────────

def print_section(title):
    sep = "─" * 65
    print(f"\n{sep}\n  {title}\n{sep}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    keys           = list(FEATURES.keys())
    feature_labels = FEATURES

    # ── 1. Load data ──────────────────────────────────────────────────────────
    print_section("Loading data")
    years, X_year, period_labels = load_year_matrix(INPUT_JSON, FEATURES, METRIC_GROUP)
    print(f"  JSON file     : {INPUT_JSON}")
    print(f"  Metric group  : {METRIC_GROUP}")
    print(f"  Year matrix   : {X_year.shape[0]} years × {X_year.shape[1]} features")
    print(f"  Years covered : {min(years)} – {max(years)}")

    _, X_period = build_period_matrix(years, X_year)
    print(f"  Period matrix : {X_period.shape[0]} periods × {X_period.shape[1]} features")

    print("\n  First 5 years / first 3 raw feature values:")
    for y, row in zip(years[:5], X_year[:5]):
        print(f"    {y}: {row[:3]}")

    # ── 2. Standardize ────────────────────────────────────────────────────────
    Xz_year,   *_ = standardize(X_year)
    Xz_period, *_ = standardize(X_period)

    # ── 3. Cronbach's alpha ───────────────────────────────────────────────────
    print_section("Cronbach's Alpha — full feature set")
    a_year   = cronbach_alpha(Xz_year)
    a_period = cronbach_alpha(Xz_period)
    print(f"  Year-level   α = {a_year:.4f}   → {interp_alpha(a_year)}")
    print(f"  Period-level α = {a_period:.4f}   → {interp_alpha(a_period)}")
    print("  (Period-level based on 4 obs. — indicative only)")

    # ── 4. Alpha if item deleted ──────────────────────────────────────────────
    print_section("Alpha if Item Deleted (year-level)")
    aid_y = alpha_if_deleted(Xz_year, keys)
    for k, v in aid_y.items():
        flag = " ← improves α" if v > a_year else ""
        print(f"  {FEATURES[k]:<35} α = {v:.4f}{flag}")

    # ── 5. Backward elimination ───────────────────────────────────────────────
    print_section("Best Subset — Backward Elimination")
    best_alpha, best_keys, df_elim = backward_elimination(Xz_year, keys, feature_labels)
    print(f"  Best α        = {best_alpha:.4f}   → {interp_alpha(best_alpha)}")
    print(f"  Best features :")
    for k in best_keys:
        print(f"    • {FEATURES[k]}")

    # ── 6. PCA — all features ─────────────────────────────────────────────────
    print_section("PCA — all features (year-level)")
    eigvals, ratio, comps = pca_numpy(Xz_year)

    print("\n  Eigenvalues:")
    for i, ev in enumerate(eigvals, 1):
        print(f"    PC{i}: {ev:.4f}   ({ratio[i-1]*100:.1f}%)  "
              + ("← keep" if ev > 1 else ""))

    print("\n  Loadings (PC1–PC3):")
    for j in range(min(3, len(eigvals))):
        print(f"\n  Component {j+1}:")
        for k, loading in zip(keys, comps[j]):
            print(f"    {FEATURES[k]:<35} {loading:>8.4f}")

    # ── PC scores (PC1–PC3) ───────────────────────────────────────────────────
    pc_scores_3 = Xz_year @ comps[:3].T        # shape: (n_years, 3)
    pc1_all     = pc_scores_3[:, 0]
    pc2_all     = pc_scores_3[:, 1]
    pc3_all     = pc_scores_3[:, 2]

    # NOTE: PCA signs are arbitrary. If you want PC1 to increase with
    # "complexity", you can flip the sign (uncomment the two lines below):
    # pc1_all             *= -1
    # pc_scores_3[:, 0]   *= -1

    print(f"\n  PC1 explains {ratio[0]*100:.1f}% of variance")
    print("\n  PC1 scores by year (first 10):")
    for y, s in zip(years[:10], pc1_all[:10]):
        print(f"    {y}: {s:.4f}")

    # ── 7. PCA — best subset ──────────────────────────────────────────────────
    print_section("PCA — best subset (year-level)")
    best_idx     = [keys.index(k) for k in best_keys]
    Xz_best      = Xz_year[:, best_idx]
    _, _, bcomps = pca_numpy(Xz_best)
    bpc1         = Xz_best @ bcomps[0]
    print(f"  Using {len(best_keys)} features for best-subset composite.")

    # ── 8. Period composites ──────────────────────────────────────────────────
    print_section("Period Composite Scores")
    df_pc1_per  = make_period_composite_df(years, pc1_all, 'Mean PC1 (all features)')
    df_bpc1_per = make_period_composite_df(years, bpc1,    'Mean PC1 (best subset)')

    print("\n  All features:")
    print(df_pc1_per.to_string(index=False))
    print("\n  Best subset:")
    print(df_bpc1_per.to_string(index=False))

    # ── 9. DataFrames for Excel ───────────────────────────────────────────────
    df_alpha    = make_alpha_summary_df(Xz_year, Xz_period, a_year, a_period)
    df_aid      = make_aid_df(Xz_year, Xz_period, a_year, a_period, keys, feature_labels)
    df_eig      = make_eigenvalue_df(eigvals, ratio)
    df_load     = make_loadings_df(comps, keys, feature_labels)
    df_pc1_yr   = make_pc1_year_df(years, pc1_all, period_labels)
    df_pc123_yr = make_pc123_year_df(years, pc_scores_3, period_labels)
    df_corr     = make_corr_df(Xz_year, keys, feature_labels)

    # ── 10. Scree plot ────────────────────────────────────────────────────────
    print_section("Scree Plot")
    try:
        plot_scree(eigvals, set(best_keys), keys)
    except Exception as e:
        print(f"  Could not show scree plot: {e}")

    # ── 11. PC trends plot (PC1–PC3) ──────────────────────────────────────────
    print_section("PC Trends Plot (PC1–PC3 over time)")
    try:
        plot_pc_trends(years, pc1_all, pc2_all, pc3_all, outpath="pc_trends.png")
    except Exception as e:
        print(f"  Could not plot PC trends: {e}")

    # ── 12. PC1 year-level plot (best subset) ─────────────────────────────────
    print_section("PC1 Year-Level Plot (best subset)")
    try:
        plot_pc1_years(years, bpc1, outpath="pc1_over_time.png")
    except Exception as e:
        print(f"  Could not plot PC1 year-level: {e}")

    # ── 13. Write Excel ───────────────────────────────────────────────────────
    print_section(f"Writing {OUTPUT_XLS}")
    wb = Workbook()
    wb.remove(wb.active)

    sheets = [
        ('1 – Cronbach Alpha',
         df_alpha,
         "Cronbach's Alpha — Full Feature Set (Year & Period Level)"),
        ('2 – Alpha If Item Deleted',
         df_aid,
         "Cronbach's Alpha — If Item Deleted (Year & Period Level)"),
        ('3 – Best Subset',
         df_elim,
         "Best Composite Subset — Backward Elimination (Year-Level)"),
        ('4 – PCA Eigenvalues',
         df_eig,
         "PCA Scree Data — Eigenvalues & Explained Variance (Year-Level)"),
        ('5 – PCA Loadings',
         df_load,
         "PCA Component Loadings (Year-Level, All Features)"),
        ('6 – PC1 by Year',
         df_pc1_yr,
         "PC1 Complexity Index Score by Year"),
        ('6b – PC1–PC3 by Year',
         df_pc123_yr,
         "PC1–PC3 Scores by Year (All Features, Year-Level PCA)"),
        ('7 – Period Composite (All)',
         df_pc1_per,
         "PC1 Complexity Composite by Period — All Features"),
        ('8 – Period Composite (Best)',
         df_bpc1_per,
         "PC1 Complexity Composite by Period — Best Subset"),
    ]

    for sheet_name, df, title in sheets:
        ws = wb.create_sheet(sheet_name)
        write_df(ws, df, title)

    ws_corr = wb.create_sheet('9 – Correlation Matrix')
    write_corr(ws_corr, df_corr,
               "Feature Correlation Matrix (Year-Level, Standardized)")

    wb.save(OUTPUT_XLS)
    print(f"\n  ✓ Saved: {OUTPUT_XLS}")
    print("  ✓ Analysis complete!\n")


if __name__ == '__main__':
    main()


import os
print(f"\n  Working directory: {os.getcwd()}")