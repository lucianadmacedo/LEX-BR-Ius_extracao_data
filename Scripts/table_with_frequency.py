#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Aggregate complexity metrics by period and export to Excel.
Also counts /modificacao_(YEAR) occurrences from annotated_tags.xml.

Outputs: complexity_periods_analysis.xlsx
"""

import json
import re
import pandas as pd
import numpy as np
from collections import OrderedDict, defaultdict
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ---- Config ----
INPUT_JSON = 'complexity_analysis_data.json'
INPUT_XML  = 'annotated_tags.xml'
OUTPUT_XLS = 'complexity_periods_analysis.xlsx'

BINS = [
    ("1943–1962", 1943, 1962),
    ("1964–1986", 1964, 1986),
    ("1988–2016", 1988, 2016),
    ("2017–2022", 2017, 2022),
]

FEATURES = OrderedDict([
    ('word_order_inversion',  'Word Order Inversion (VS)'),
    ('nominalization',        'Nominalizations'),
    ('subordination',         'Subordinate Clauses'),
    ('passive',               'Passive Constructions'),
    ('gerunds_participles',   'Gerunds & Participles'),
    ('relative_pronouns',     'Relative Pronouns & Conj.'),
    ('appositions',           'Appositions & Parentheticals'),
    ('prepositional_phrases', 'Prepositional Phrases'),
    ('complex_verb_forms',    'Complex Verb Forms'),
])


# ============================================================
# 1. Count /modificacao_(YEAR) in the XML file
# ============================================================

def count_modificacoes_from_xml(xml_path):
    """
    Reads annotated_tags.xml as plain text and counts all
    occurrences of /modificacao_YYYY or \modificacao_YYYY
    (handles both slash types).
    Returns dict: {year: count}
    """
    year_counts = defaultdict(int)

    # Pattern: forward or back slash, then 'modificacao_', then 4 digits
    pattern = re.compile(r'[/\\]modificacao_(\d{4})', re.IGNORECASE)

    print(f"Reading {xml_path} ...")
    with open(xml_path, 'r', encoding='utf-8', errors='replace') as f:
        for line in f:
            for match in pattern.finditer(line):
                year = int(match.group(1))
                year_counts[year] += 1

    print(f"  Found references for {len(year_counts)} distinct years.")
    return dict(sorted(year_counts.items()))


def assign_period(year, bins):
    for label, start, end in bins:
        if start <= year <= end:
            return label
    return None


def build_modificacoes_tables(year_counts, bins):
    """
    Returns two DataFrames:
    1. Per‑year counts
    2. Per‑period totals
    """
    # Per year
    rows_year = []
    for year in sorted(year_counts.keys()):
        rows_year.append({
            'Year': year,
            'Period': assign_period(year, bins) or 'Outside bins',
            'Modificações (count)': year_counts[year],
        })
    df_year = pd.DataFrame(rows_year)

    # Per period
    rows_period = []
    for label, start, end in bins:
        total = sum(
            v for y, v in year_counts.items() if start <= y <= end
        )
        n_years = sum(1 for y in year_counts if start <= y <= end)
        rows_period.append({
            'Period': label,
            'Years with data': n_years,
            'Total modificações': total,
        })
    df_period = pd.DataFrame(rows_period)

    return df_year, df_period


# ============================================================
# 2. Aggregate JSON into periods
# ============================================================

def load_json(path):
    with open(path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return sorted(((int(y), rec) for y, rec in data.items()), key=lambda x: x[0])


def aggregate_periods(year_items, bins, features):
    periods = [label for (label, _, _) in bins]
    summary = {}

    for label, start, end in bins:
        items = [(y, r) for (y, r) in year_items if start <= y <= end]
        if not items:
            continue

        total_files = total_sents = total_tokens = total_types = 0
        feat_sum_ps  = {k: 0.0 for k in features.keys()}
        feat_sum_p100 = {k: 0.0 for k in features.keys()}

        for year, rec in items:
            sents = rec['sentences']
            toks  = rec['tokens']

            total_files  += rec['files']
            total_sents  += sents
            total_tokens += toks
            total_types  += rec.get('types', 0)

            ps   = rec.get('per_sentence', {})
            p100 = rec.get('per_100_tokens', {})

            for feat in features.keys():
                if feat in ps:
                    feat_sum_ps[feat]   += ps[feat]   * sents
                if feat in p100:
                    feat_sum_p100[feat] += p100[feat] * toks

        ttr = total_types / total_tokens if total_tokens > 0 else 0.0

        ps_means   = {f: feat_sum_ps[f]   / total_sents   if total_sents   > 0 else 0.0 for f in features}
        p100_means = {f: feat_sum_p100[f] / total_tokens  if total_tokens  > 0 else 0.0 for f in features}

        summary[label] = {
            'files':      total_files,
            'sentences':  total_sents,
            'tokens':     total_tokens,
            'types':      total_types,
            'ttr':        ttr,
            'per_sentence':   ps_means,
            'per_100_tokens': p100_means,
        }

    return periods, summary


# ============================================================
# 3. Build DataFrames
# ============================================================

def build_period_stats_df(periods, summary):
    rows = []
    for label in periods:
        if label not in summary:
            continue
        s = summary[label]
        rows.append({
            'Period':            label,
            'Files (Modif.)':    s['files'],
            'Sentences':         s['sentences'],
            'Tokens':            s['tokens'],
            'Types (approx.)':   s['types'],
            'TTR (approx.)':     round(s['ttr'], 4),
        })
    return pd.DataFrame(rows)


def build_feature_df(periods, summary, metric_group, features):
    """
    Rows = features, Columns = periods.
    """
    rows = []
    for feat_key, feat_label in features.items():
        row = {'Feature': feat_label}
        for label in periods:
            if label not in summary:
                row[label] = None
            else:
                row[label] = round(summary[label][metric_group][feat_key], 4)
        rows.append(row)
    return pd.DataFrame(rows)


# ============================================================
# 4. Write Excel with formatting
# ============================================================

HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
PERIOD_FILL   = PatternFill("solid", fgColor="2E75B6")   # medium blue
ALT_ROW_FILL  = PatternFill("solid", fgColor="DCE6F1")   # light blue
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
PERIOD_FONT   = Font(bold=True, color="FFFFFF", size=10)
NORMAL_FONT   = Font(size=10)
CENTER        = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT          = Alignment(horizontal='left',   vertical='center', wrap_text=True)
THIN          = Side(style='thin', color='BFBFBF')
THIN_BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def format_sheet(ws, df, title, index=False):
    """
    Write a DataFrame to a worksheet with formatting.
    """
    # Title row
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=len(df.columns) + (1 if index else 0))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font      = Font(bold=True, color="FFFFFF", size=13)
    title_cell.fill      = HEADER_FILL
    title_cell.alignment = CENTER

    # Write DataFrame (headers at row 2)
    col_offset = 2 if index else 1

    # Header row (row 2)
    header_cols = (list(df.index.names) if index else []) + list(df.columns)
    for c, h in enumerate(header_cols, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font      = PERIOD_FONT
        cell.fill      = PERIOD_FILL
        cell.alignment = CENTER
        cell.border    = THIN_BORDER

    # Data rows
    for r_idx, (_, row) in enumerate(df.iterrows(), start=3):
        fill = ALT_ROW_FILL if r_idx % 2 == 0 else WHITE_FILL
        values = list(row.index) if index else []
        values += list(row.values)
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font      = NORMAL_FONT
            cell.fill      = fill
            cell.alignment = LEFT if c_idx == 1 else CENTER
            cell.border    = THIN_BORDER

    # Auto-size columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Freeze header rows
    ws.freeze_panes = 'A3'


def write_excel(
    output_path,
    df_stats,
    df_feat_ps,
    df_feat_p100,
    df_modif_year,
    df_modif_period,
):
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # --- Sheet 1: Period statistics ---
        df_stats.to_excel(writer, sheet_name='Period Statistics', index=False, startrow=1)
        ws = writer.sheets['Period Statistics']
        format_sheet(ws, df_stats, 'Period Statistics: Files, Tokens, Types, TTR')

        # --- Sheet 2: Features per sentence ---
        df_feat_ps.to_excel(writer, sheet_name='Features per Sentence', index=False, startrow=1)
        ws = writer.sheets['Features per Sentence']
        format_sheet(ws, df_feat_ps, 'Complexity Features — Mean per Sentence')

        # --- Sheet 3: Features per 100 tokens ---
        df_feat_p100.to_excel(writer, sheet_name='Features per 100 Tokens', index=False, startrow=1)
        ws = writer.sheets['Features per 100 Tokens']
        format_sheet(ws, df_feat_p100, 'Complexity Features — Mean per 100 Tokens')

        # --- Sheet 4: Modificações per year ---
        df_modif_year.to_excel(writer, sheet_name='Modificações per Year', index=False, startrow=1)
        ws = writer.sheets['Modificações per Year']
        format_sheet(ws, df_modif_year, 'Modificações per Year (from annotated_tags.xml)')

        # --- Sheet 5: Modificações per period ---
        df_modif_period.to_excel(writer, sheet_name='Modificações per Period', index=False, startrow=1)
        ws = writer.sheets['Modificações per Period']
        format_sheet(ws, df_modif_period, 'Modificações per Period (from annotated_tags.xml)')

    print(f"\n✓ Excel file written to: {output_path}")


# ============================================================
# 5. Main
# ============================================================

def main():
    # Load and aggregate JSON
    print("Loading complexity_analysis_data.json ...")
    year_items = load_json(INPUT_JSON)
    periods, summary = aggregate_periods(year_items, BINS, FEATURES)

    # Build DataFrames
    df_stats     = build_period_stats_df(periods, summary)
    df_feat_ps   = build_feature_df(periods, summary, 'per_sentence',   FEATURES)
    df_feat_p100 = build_feature_df(periods, summary, 'per_100_tokens', FEATURES)

    # Count modificações from XML
    year_counts = count_modificacoes_from_xml(INPUT_XML)
    df_modif_year, df_modif_period = build_modificacoes_tables(year_counts, BINS)

    # Print summaries to console
    print("\n--- Period Statistics ---")
    print(df_stats.to_string(index=False))

    print("\n--- Modificações per Period ---")
    print(df_modif_period.to_string(index=False))

    print("\n--- Features per Sentence ---")
    print(df_feat_ps.to_string(index=False))

    # Write to Excel
    write_excel(
        OUTPUT_XLS,
        df_stats,
        df_feat_ps,
        df_feat_p100,
        df_modif_year,
        df_modif_period,
    )


if __name__ == '__main__':
    main()